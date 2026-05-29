import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import date
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import backend

# Inicializador del FrontEnd, con un Try validador.

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False


# Clase App, base visual.
class App(ttk.Window):

    def __init__(self) -> None:
        super().__init__(themename="darkly")
        self.title("Control de Gastos")
        self.geometry("1000x660")
        self.minsize(800, 600)
        self.resizable(True, True)   

        self._init_db()
        self._init_estado()
        self._build_layout()
        self._build_sidebar()
        self._build_hoy()
        self._build_mensual()
        self._build_categorias()
        self._arrancar()
    
    def _init_db(self) -> None:
        self.db = backend.GastosDB()
        self.db.inicializar_db()


    def _init_estado(self) -> None:
        self.temas = ["darkly", "flatly"]
        self.tema_actual = 0
        self.ultimos_datos_grafico = [0, 0]

    def _build_layout(self) -> None:
        frame_principal = ttk.Frame(self)
        frame_principal.pack(fill="both", expand=True)

        self.sidebar = ttk.Frame(frame_principal, width=180, bootstyle="dark")
        self.sidebar.pack(side="left", fill="y")
        self.sidebar.pack_propagate(False)

        self.contenido = ttk.Frame(frame_principal)
        self.contenido.pack(side="left", fill="both", expand=True, padx=10, pady=10)
        self.contenido.rowconfigure(0, weight=1)
        self.contenido.columnconfigure(0, weight=1)

        self.frame_hoy = ttk.Frame(self.contenido)
        self.frame_mes = ttk.Frame(self.contenido)
        self.frame_cat = ttk.Frame(self.contenido)

        for frame in (self.frame_hoy, self.frame_mes, self.frame_cat):
            frame.place(relwidth=1, relheight=1)

    def _build_sidebar(self) -> None:
        ttk.Label(
            self.sidebar, text="💰",
            font=("Arial", 28),
            bootstyle="inverse-dark"
        ).pack(pady=(20, 0))

        ttk.Label(
            self.sidebar, text="Control de\nGastos",
            font=("Arial", 11, "bold"),
            bootstyle="inverse-dark",
            justify="center"
        ).pack(pady=(4, 20))

        ttk.Separator(self.sidebar, orient="horizontal").pack(fill="x", padx=12, pady=4)

        self.botones_nav = []

        self.btn_hoy = ttk.Button(self.sidebar, text="🏠   Hoy",
            bootstyle="dark", width=18,
            command=lambda: [self.mostrar(self.frame_hoy, self.btn_hoy),
                            self.cargar_gastos_hoy()])

        self.btn_mes = ttk.Button(self.sidebar, text="📊   Mensual",
            bootstyle="dark", width=18,
            command=lambda: [self.mostrar(self.frame_mes, self.btn_mes),
                            self.cargar_gastos_mes()])

        self.btn_cat = ttk.Button(self.sidebar, text="🏷️   Categorías",
            bootstyle="dark", width=18,
            command=lambda: [self.mostrar(self.frame_cat, self.btn_cat),
                            self.cargar_categorias()])

        for btn in (self.btn_hoy, self.btn_mes, self.btn_cat):
            self.botones_nav.append(btn)
            btn.pack(pady=3, padx=10, fill="x")

        ttk.Separator(self.sidebar, orient="horizontal").pack(fill="x", padx=12, pady=12)

        ttk.Button(
            self.sidebar, text="🌓   Tema",
            bootstyle="secondary", width=18,
            command=self.alternar_tema
        ).pack(side="bottom", pady=12, padx=10, fill="x")

    def _build_hoy(self) -> None:
        form_hoy = ttk.Frame(self.frame_hoy)
        form_hoy.pack(fill="x", pady=(0, 10))
        form_hoy.columnconfigure(1, weight=1)
        form_hoy.columnconfigure(3, weight=1)

        ttk.Label(form_hoy, text="Monto ($)").grid(row=0, column=0, sticky="e", padx=8, pady=6)
        self.entry_monto = ttk.Entry(form_hoy, width=22)
        self.entry_monto.grid(row=0, column=1, padx=8, pady=6)

        ttk.Label(form_hoy, text="Categoría").grid(row=0, column=2, sticky="e", padx=8, pady=6)
        self.combo_categoria = ttk.Combobox(form_hoy, width=20,
                                            values=self.db.validador.CATEGORIAS_DEFAULT)
        self.combo_categoria.grid(row=0, column=3, padx=8, pady=6)
        self.combo_categoria.current(0)

        ttk.Label(form_hoy, text="Descripción").grid(row=1, column=0, sticky="e", padx=8, pady=6)
        self.entry_descripcion = ttk.Entry(form_hoy, width=22)
        self.entry_descripcion.grid(row=1, column=1, padx=8, pady=6)

        ttk.Label(form_hoy, text="Tipo").grid(row=1, column=2, sticky="e", padx=8, pady=6)
        self.combo_tipo = ttk.Combobox(form_hoy, width=20, state="readonly",
                                    values=["Personal", "Negocio"])
        self.combo_tipo.grid(row=1, column=3, padx=8, pady=6)
        self.combo_tipo.current(0)

        ttk.Button(
            form_hoy, text="Registrar Gasto",
            bootstyle="success", width=20,
            command=self.registrar_gasto
        ).grid(row=2, column=0, columnspan=4, pady=12)

        tabla_frame_hoy = ttk.Frame(self.frame_hoy)
        tabla_frame_hoy.pack(fill="both", expand=True)

        cols = ("fecha", "descripcion", "categoria", "monto", "tipo")
        self.tabla_hoy = ttk.Treeview(tabla_frame_hoy, columns=cols,
                                   show="headings", height=10, bootstyle="dark")

        self.tabla_hoy.heading("fecha",       text="Fecha")
        self.tabla_hoy.heading("descripcion", text="Descripción")
        self.tabla_hoy.heading("categoria",   text="Categoría")
        self.tabla_hoy.heading("monto",       text="Monto")
        self.tabla_hoy.heading("tipo",        text="Tipo")

        self.tabla_hoy.column("fecha",       width=95,  anchor="center")
        self.tabla_hoy.column("descripcion", width=250)
        self.tabla_hoy.column("categoria",   width=150)
        self.tabla_hoy.column("monto",       width=110, anchor="e")
        self.tabla_hoy.column("tipo",        width=95,  anchor="center")

        scroll_hoy = ttk.Scrollbar(tabla_frame_hoy, orient="vertical",
                                    command=self.tabla_hoy.yview)
        self.tabla_hoy.configure(yscrollcommand=scroll_hoy.set)
        self.tabla_hoy.pack(side="left", fill="both", expand=True)
        scroll_hoy.pack(side="right", fill="y")

        acciones_hoy = ttk.Frame(self.frame_hoy)
        acciones_hoy.pack(fill="x", pady=(8, 0))

        self.label_total_hoy = ttk.Label(acciones_hoy, text="Total del día: $0",
                                        font=("Arial", 12, "bold"), bootstyle="info")
        self.label_total_hoy.pack(side="left", padx=4)

        ttk.Button(acciones_hoy, text="🗑️  Eliminar",
            bootstyle="danger-outline", width=14,
            command=self.eliminar_gasto_seleccionado
        ).pack(side="right", padx=4)

        ttk.Button(acciones_hoy, text="✏️  Editar",
            bootstyle="warning-outline", width=14,
            command=lambda: self.abrir_ventana_editar(self.tabla_hoy)
        ).pack(side="right", padx=4)

    def _build_mensual(self) -> None:
        selector_mes = ttk.Frame(self.frame_mes)
        selector_mes.pack(fill="x", pady=(0, 10))
        selector_mes.columnconfigure(1, weight=1)

        ttk.Label(selector_mes, text="Mes:").pack(side="left", padx=(0, 4))
        self.combo_mes = ttk.Combobox(selector_mes, width=14, state="readonly", values=[
            "1 - Enero", "2 - Febrero", "3 - Marzo", "4 - Abril",
            "5 - Mayo", "6 - Junio", "7 - Julio", "8 - Agosto",
            "9 - Septiembre", "10 - Octubre", "11 - Noviembre", "12 - Diciembre"
        ])
        self.combo_mes.pack(side="left", padx=(0, 12))
        self.combo_mes.current(date.today().month - 1)

        ttk.Label(selector_mes, text="Año:").pack(side="left", padx=(0, 4))
        año_min = self.db.obtener_año_minimo()
        self.combo_año = ttk.Combobox(selector_mes, width=8, state="readonly",
                                    values=[str(y) for y in range(año_min, date.today().year + 2)])
        self.combo_año.pack(side="left", padx=(0, 12))
        self.combo_año.set(str(date.today().year))

        ttk.Button(selector_mes, text="🔍  Buscar",
            bootstyle="primary", command=self.cargar_gastos_mes
        ).pack(side="left")

        tabla_frame_mes = ttk.Frame(self.frame_mes)
        tabla_frame_mes.pack(fill="both", expand=True)

        cols = ("fecha", "descripcion", "categoria", "monto", "tipo")
        self.tabla_mes = ttk.Treeview(tabla_frame_mes, columns=cols,
                                    show="headings", height=13, bootstyle="dark")

        self.tabla_mes.heading("fecha",       text="Fecha")
        self.tabla_mes.heading("descripcion", text="Descripción")
        self.tabla_mes.heading("categoria",   text="Categoría")
        self.tabla_mes.heading("monto",       text="Monto")
        self.tabla_mes.heading("tipo",        text="Tipo")

        self.tabla_mes.column("fecha",       width=95,  anchor="center")
        self.tabla_mes.column("descripcion", width=250)
        self.tabla_mes.column("categoria",   width=150)
        self.tabla_mes.column("monto",       width=110, anchor="e")
        self.tabla_mes.column("tipo",        width=95,  anchor="center")

        scroll_mes = ttk.Scrollbar(tabla_frame_mes, orient="vertical",
                                command=self.tabla_mes.yview)
        self.tabla_mes.configure(yscrollcommand=scroll_mes.set)
        self.tabla_mes.pack(side="left", fill="both", expand=True)
        scroll_mes.pack(side="right", fill="y")

        resumen_mes = ttk.Frame(self.frame_mes)
        resumen_mes.pack(fill="x", pady=(8, 0))

        self.label_total_mes = ttk.Label(resumen_mes, text="Total: $0",
                                         font=("Arial", 12, "bold"), bootstyle="info")
        self.label_total_personal = ttk.Label(resumen_mes, text="Personal: $0",
                                              font=("Arial", 11), bootstyle="primary")
        self.label_total_negocio = ttk.Label(resumen_mes, text="Negocio: $0",
                                             font=("Arial", 11), bootstyle="success")

        self.label_total_mes.pack(side="left", padx=10)
        self.label_total_personal.pack(side="left", padx=10)
        self.label_total_negocio.pack(side="left", padx=10)

        ttk.Button(resumen_mes, text="📥  Exportar",
            bootstyle="secondary", command=self.exportar_mes
        ).pack(side="right", padx=8)

        ttk.Button(resumen_mes, text="🗑️  Eliminar",
            bootstyle="danger-outline", width=14,
            command=self.eliminar_gasto_mes
        ).pack(side="right", padx=4)

        ttk.Button(resumen_mes, text="✏️  Editar",
            bootstyle="warning-outline", width=14,
            command=lambda: self.abrir_ventana_editar(self.tabla_mes)
        ).pack(side="right", padx=4)

    def _build_categorias(self) -> None:
        selector_cat = ttk.Frame(self.frame_cat)
        selector_cat.pack(fill="x", pady=(0, 10))
        selector_cat.columnconfigure(1, weight=1)

        ttk.Label(selector_cat, text="Mes:").pack(side="left", padx=(0, 4))
        self.combo_mes_cat = ttk.Combobox(selector_cat, width=14, state="readonly", values=[
            "1 - Enero", "2 - Febrero", "3 - Marzo", "4 - Abril",
            "5 - Mayo", "6 - Junio", "7 - Julio", "8 - Agosto",
            "9 - Septiembre", "10 - Octubre", "11 - Noviembre", "12 - Diciembre"
        ])
        self.combo_mes_cat.pack(side="left", padx=(0, 12))
        self.combo_mes_cat.current(date.today().month - 1)

        ttk.Label(selector_cat, text="Año:").pack(side="left", padx=(0, 4))
        año_min = self.db.obtener_año_minimo()
        self.combo_año_cat = ttk.Combobox(selector_cat, width=8, state="readonly",
                                        values=[str(y) for y in range(año_min, date.today().year + 2)])
        self.combo_año_cat.pack(side="left", padx=(0, 12))
        self.combo_año_cat.set(str(date.today().year))

        ttk.Button(selector_cat, text="🔍  Buscar",
            bootstyle="primary", command=self.cargar_categorias
        ).pack(side="left")

        tabla_frame_cat = ttk.Frame(self.frame_cat)
        tabla_frame_cat.pack(fill="both", expand=True)

        self.tabla_cat = ttk.Treeview(tabla_frame_cat,
            columns=("categoria", "total"), show="headings",
            height=13, bootstyle="dark")

        self.tabla_cat.heading("categoria", text="Categoría")
        self.tabla_cat.heading("total",     text="Total gastado")
        self.tabla_cat.column("categoria",  width=420)
        self.tabla_cat.column("total",      width=200, anchor="e")

        scroll_cat = ttk.Scrollbar(tabla_frame_cat, orient="vertical",
                                    command=self.tabla_cat.yview)
        self.tabla_cat.configure(yscrollcommand=scroll_cat.set)
        self.tabla_cat.pack(side="left", fill="both", expand=True)
        scroll_cat.pack(side="right", fill="y")

        self.label_total_cat = ttk.Label(self.frame_cat, text="Total: $0",
                                     font=("Arial", 12, "bold"), bootstyle="info")
        self.label_total_cat.pack(pady=(8, 4))

        self.canvas_grafico = tk.Canvas(self.frame_cat, height=80, bg="#222222",
                                    highlightthickness=0)
        self.canvas_grafico.pack(fill="x", padx=4, pady=(0, 4))

    def _arrancar(self) -> None:
        self.mostrar(self.frame_hoy, self.btn_hoy)
        self.cargar_gastos_hoy()

    def mostrar(self, frame, btn_activo) -> None:
        frame.tkraise()
        for btn in self.botones_nav:
            btn.configure(bootstyle="dark")
        btn_activo.configure(bootstyle="primary")

    def alternar_tema(self) -> None:
        self.tema_actual = 1 - self.tema_actual
        self.style.theme_use(self.temas[self.tema_actual])
        self.dibujar_grafico(self.ultimos_datos_grafico[0],
                             self.ultimos_datos_grafico[1])

    def registrar_gasto(self) -> None:
        exito, mensaje = self.db.agregar_gasto(
            self.entry_descripcion.get(),
            self.combo_categoria.get(),
            self.entry_monto.get(),
            self.combo_tipo.get()
        )
        if exito:
            messagebox.showinfo("Éxito", mensaje)
            self.entry_monto.delete(0, "end")
            self.entry_descripcion.delete(0, "end")
            self.combo_categoria.current(0)
            self.combo_tipo.current(0)
            self.cargar_gastos_hoy()
            self.cargar_gastos_mes()
        else:
            messagebox.showerror("Error", mensaje)

    def cargar_gastos_hoy(self) -> None:
        for item in self.tabla_hoy.get_children():
            self.tabla_hoy.delete(item)

        exito, resultado = self.db.obtener_gastos_dia()
        total_dia = 0

        if exito:
            for fila in resultado["data"]:
                id_gasto, fecha, descripcion, categoria, monto, tipo = fila
                total_dia += monto
                self.tabla_hoy.insert("", "end",
                    values=(fecha, descripcion, categoria, f"${monto:,.0f}", tipo),
                    tags=(str(id_gasto),))
            self.label_total_hoy.config(
                text=f"Total del día: ${total_dia:,.0f}  ({resultado['count']} gastos)"
            )
        else:
            messagebox.showerror("Error", f"No se pudieron cargar los gastos: {resultado}")

    def eliminar_gasto_seleccionado(self) -> None:
        seleccion = self.tabla_hoy.selection()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Selecciona un gasto para eliminar.")
            return
        if not messagebox.askyesno("Confirmar", "¿Estás seguro de que quieres eliminar este gasto?"):
            return
        id_gasto = int(self.tabla_hoy.item(seleccion[0], "tags")[0])
        exito, mensaje = self.db.eliminar_gasto(id_gasto)
        if exito:
            self.cargar_gastos_hoy()
            self.cargar_gastos_mes()
        else:
            messagebox.showerror("Error", mensaje)

    def eliminar_gasto_mes(self) -> None:
        seleccion = self.tabla_mes.selection()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Selecciona un gasto para eliminar.")
            return
        if not messagebox.askyesno("Confirmar", "¿Estás seguro de que quieres eliminar este gasto?"):
            return
        id_gasto = int(self.tabla_mes.item(seleccion[0], "tags")[0])
        exito, mensaje = self.db.eliminar_gasto(id_gasto)
        if exito:
            self.cargar_gastos_mes()
        else:
            messagebox.showerror("Error", mensaje)

    def abrir_ventana_editar(self, tabla=None) -> None:
        if tabla is None:
            tabla = self.tabla_hoy
        seleccion = tabla.selection()
        if not seleccion:
            messagebox.showwarning("Advertencia", "Selecciona un gasto para editar.")
            return

        item     = seleccion[0]
        id_gasto = int(tabla.item(item, "tags")[0])
        valores  = tabla.item(item, "values")

        ventana = ttk.Toplevel(self)
        ventana.title("Editar Gasto")
        ventana.geometry("400x340")
        ventana.resizable(False, False)
        ventana.grab_set()

        ttk.Label(ventana, text="Editar Gasto",
                  font=("Arial", 14, "bold")).pack(pady=(18, 10))

        frame = ttk.Frame(ventana, padding=10)
        frame.pack()

        monto_raw = valores[3].replace("$", "").replace("\xa0", "")
        try:
            monto_para_editar = str(round(backend._parsear_monto(monto_raw), 2))
        except (ValueError, KeyError):
            monto_para_editar = monto_raw

        campos = [
            ("Fecha (YYYY-MM-DD):", valores[0]),
            ("Descripción:",        valores[1]),
            ("Monto ($):",          monto_para_editar),
        ]
        entries = []
        for i, (label, valor) in enumerate(campos):
            ttk.Label(frame, text=label).grid(row=i, column=0, sticky="e", pady=5, padx=6)
            e = ttk.Entry(frame, width=24)
            e.insert(0, valor)
            e.grid(row=i, column=1, pady=5, padx=6)
            entries.append(e)

        e_fecha, e_desc, e_monto = entries

        ttk.Label(frame, text="Categoría:").grid(row=3, column=0, sticky="e", pady=5, padx=6)
        e_cat = ttk.Combobox(frame, width=22, values=self.db.validador.CATEGORIAS_DEFAULT)
        e_cat.set(valores[2])
        e_cat.grid(row=3, column=1, pady=5, padx=6)

        ttk.Label(frame, text="Tipo:").grid(row=4, column=0, sticky="e", pady=5, padx=6)
        e_tipo = ttk.Combobox(frame, width=22, state="readonly", values=["Personal", "Negocio"])
        e_tipo.set(valores[4])
        e_tipo.grid(row=4, column=1, pady=5, padx=6)

        def guardar_edicion() -> None:
            exito, mensaje = self.db.editar_gasto(
                id_gasto,
                e_fecha.get(), e_cat.get(), e_desc.get(), e_monto.get(), e_tipo.get()
            )
            if exito:
                messagebox.showinfo("Éxito", mensaje, parent=ventana)
                ventana.destroy()
                self.cargar_gastos_hoy()
                self.cargar_gastos_mes()
            else:
                messagebox.showerror("Error", mensaje, parent=ventana)

        ttk.Button(ventana, text="💾  Guardar cambios",
                   bootstyle="success", command=guardar_edicion).pack(pady=14)

    def cargar_gastos_mes(self) -> None:
        for item in self.tabla_mes.get_children():
            self.tabla_mes.delete(item)

        mes = int(self.combo_mes.get().split(" ")[0])
        año = int(self.combo_año.get())

        exito, resultado = self.db.obtener_gastos_mes(mes, año)
        if not exito:
            messagebox.showerror("Error", resultado)
            return

        total = total_personal = total_negocio = 0
        for fila in resultado["data"]:
            id_gasto, fecha, descripcion, categoria, monto, tipo = fila
            total += monto
            if tipo == "Personal":
                total_personal += monto
            else:
                total_negocio += monto
            self.tabla_mes.insert("", "end",
                values=(fecha, descripcion, categoria, f"${monto:,.0f}", tipo),
                tags=(str(id_gasto),))

        self.label_total_mes.config(text=f"Total: ${total:,.0f}")
        self.label_total_personal.config(text=f"Personal: ${total_personal:,.0f}")
        self.label_total_negocio.config(text=f"Negocio: ${total_negocio:,.0f}")

    def exportar_mes(self) -> None:
        mes = int(self.combo_mes.get().split(" ")[0])
        año = int(self.combo_año.get())
        nombre_mes = self.combo_mes.get().split(" - ")[1]

        exito, resultado = self.db.obtener_gastos_mes(mes, año)
        if not exito:
            messagebox.showerror("Error", resultado)
            return
        if not resultado["data"]:
            messagebox.showwarning("Sin datos", "No hay gastos registrados en ese mes.")
            return

        ruta = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("CSV", "*.csv")],
            initialfile=f"gastos_{nombre_mes}_{año}"
        )
        if not ruta:
            return

        if ruta.endswith(".csv"):
            import csv
            with open(ruta, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["Fecha", "Descripción", "Categoría", "Monto", "Tipo"])
                for fila in resultado["data"]:
                    _, fecha, descripcion, categoria, monto, tipo = fila
                    writer.writerow([fecha, descripcion, categoria, int(monto), tipo])
        else:
            if not OPENPYXL_DISPONIBLE:
                messagebox.showerror("Librería faltante",
                    "Para exportar a Excel instala openpyxl:\n\npip install openpyxl")
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"{nombre_mes} {año}"

            encabezados = ["Fecha", "Descripción", "Categoría", "Monto", "Tipo"]
            for col, titulo in enumerate(encabezados, start=1):
                celda = ws.cell(row=1, column=col, value=titulo)
                celda.font = Font(bold=True, color="FFFFFF")
                celda.fill = PatternFill("solid", fgColor="4A90D9")
                celda.alignment = Alignment(horizontal="center")

            for fila_idx, fila in enumerate(resultado["data"], start=2):
                _, fecha, descripcion, categoria, monto, tipo = fila
                ws.cell(row=fila_idx, column=1, value=fecha)
                ws.cell(row=fila_idx, column=2, value=descripcion)
                ws.cell(row=fila_idx, column=3, value=categoria)
                ws.cell(row=fila_idx, column=4, value=monto)
                ws.cell(row=fila_idx, column=5, value=tipo)

            anchos = [12, 35, 20, 15, 12]
            for col, ancho in enumerate(anchos, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = ancho

            wb.save(ruta)

        messagebox.showinfo("Éxito", f"Archivo guardado en:\n{ruta}")

    def cargar_categorias(self) -> None:
        for item in self.tabla_cat.get_children():
            self.tabla_cat.delete(item)

        mes = int(self.combo_mes_cat.get().split(" ")[0])
        año = int(self.combo_año_cat.get())

        exito, resultado = self.db.obtener_por_categoria(mes, año)
        if not exito:
            messagebox.showerror("Error", resultado)
            return

        totales_cat = {}
        total_personal = 0
        total_negocio  = 0

        for categoria, tipo, monto in resultado["data"]:
            totales_cat[categoria] = totales_cat.get(categoria, 0) + monto
            if tipo == "Personal":
                total_personal += monto
            else:
                total_negocio += monto

        total = total_personal + total_negocio
        for categoria, monto in sorted(totales_cat.items(), key=lambda x: x[1], reverse=True):
            self.tabla_cat.insert("", "end", values=(categoria, f"${monto:,.0f}"))

        self.label_total_cat.config(text=f"Total: ${total:,.0f}")
        self.dibujar_grafico(total_personal, total_negocio)

    def dibujar_grafico(self, total_personal: float, total_negocio: float) -> None:
        self.ultimos_datos_grafico[0] = total_personal
        self.ultimos_datos_grafico[1] = total_negocio
        self.canvas_grafico.delete("all")
        total = total_personal + total_negocio

        bg = "#222222" if self.temas[self.tema_actual] == "darkly" else "#f0f0f0"
        self.canvas_grafico.configure(bg=bg)

        if total == 0:
            self.canvas_grafico.create_text(
                400, 40, text="Sin datos para mostrar",
                font=("Arial", 11), fill="#888888"
            )
            return

        ANCHO_BARRA = 680
        X_INICIO = 30
        Y_INICIO = 14
        Y_FIN    = 48

        ancho_personal = int(ANCHO_BARRA * (total_personal / total))
        ancho_negocio  = ANCHO_BARRA - ancho_personal

        self.canvas_grafico.create_rectangle(
            X_INICIO, Y_INICIO,
            X_INICIO + ancho_personal, Y_FIN,
            fill="#375a7f", outline=""
        )
        self.canvas_grafico.create_rectangle(
            X_INICIO + ancho_personal, Y_INICIO,
            X_INICIO + ancho_personal + ancho_negocio, Y_FIN,
            fill="#00bc8c", outline=""
        )

        pct_personal = (total_personal / total) * 100
        pct_negocio  = (total_negocio  / total) * 100

        self.canvas_grafico.create_rectangle(150, 54, 164, 66, fill="#375a7f", outline="")
        self.canvas_grafico.create_text(230, 60,
            text=f"Personal: {pct_personal:.1f}%", font=("Arial", 10), fill="#aaaaaa")

        self.canvas_grafico.create_rectangle(380, 54, 394, 66, fill="#00bc8c", outline="")
        self.canvas_grafico.create_text(460, 60,
            text=f"Negocio: {pct_negocio:.1f}%", font=("Arial", 10), fill="#aaaaaa")



if __name__ == "__main__":
    app = App()
    app.mainloop()

